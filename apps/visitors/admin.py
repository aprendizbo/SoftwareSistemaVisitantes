import csv
from datetime import timedelta
import pytz

from django.contrib import admin
from django.http import HttpResponse
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font

from .models import Visitor, Visit


# =========================================================
# CONVERSIÓN DE HORA A BOGOTÁ
# =========================================================
def get_bogota_time(dt):
    if not dt:
        return None

    if timezone.is_naive(dt):
        dt = timezone.make_aware(
            dt,
            timezone.utc
        )

    return dt.astimezone(
        pytz.timezone("America/Bogota")
    )


# =========================================================
# CSV - HISTORIAL DE VISITAS
# =========================================================
@admin.action(
    description="Extraer Historial Seleccionado (Archivo CSV)"
)
def extraer_historial_csv(
    modeladmin,
    request,
    queryset
):
    response = HttpResponse(
        content_type="text/csv; charset=utf-8-sig"
    )

    response["Content-Disposition"] = (
        'attachment; filename="historial_boccherini.csv"'
    )

    writer = csv.writer(
        response,
        delimiter=";"
    )

    writer.writerow([
        "ID REGISTRO",
        "NOMBRES",
        "APELLIDOS",
        "TIPO DOC",
        "NRO DOCUMENTO",
        "PERFIL VISITANTE",
        "EMPRESA CORPORATIVA",
        "PERSONA A VISITAR",
        "FECHA HORA INGRESO",
        "FECHA HORA SALIDA",
        "ESTADO ACTUAL",
    ])

    for visita in queryset.select_related("visitor"):

        entrada_local = get_bogota_time(
            visita.entry_time
        )

        salida_local = get_bogota_time(
            getattr(
                visita,
                "exit_time",
                None
            )
        )

        visitante = visita.visitor

        tipo_doc = {
            "cedula": "CC",
            "cedula_ciudadania": "CC",
            "ce": "CE",
            "cedula_extranjeria": "CE",
            "pasaporte": "PAS",
        }.get(
            str(
                getattr(
                    visitante,
                    "document_type",
                    ""
                )
            ).lower(),
            ""
        )

        writer.writerow([
            visita.id,

            getattr(
                visitante,
                "first_name",
                ""
            ),

            getattr(
                visitante,
                "last_name",
                ""
            ),

            tipo_doc,

            getattr(
                visitante,
                "document_id",
                ""
            ),

            (
                visitante.get_visitor_type_display()
                if visitante
                else ""
            ),

            getattr(
                visitante,
                "company",
                "Particular"
            ),

            getattr(
                visita,
                "person_to_visit",
                "No Asignado"
            ),

            (
                entrada_local.strftime(
                    "%Y-%m-%d %H:%M"
                )
                if entrada_local
                else ""
            ),

            (
                salida_local.strftime(
                    "%Y-%m-%d %H:%M"
                )
                if salida_local
                else "En Instalaciones"
            ),

            (
                visita.get_status_display()
                if hasattr(
                    visita,
                    "get_status_display"
                )
                else getattr(
                    visita,
                    "status",
                    ""
                )
            ),
        ])

    return response


# =========================================================
# EXCEL - VISITANTES SELECCIONADOS
# =========================================================
@admin.action(
    description="Extraer Historial Seleccionado (Excel con Fotos)"
)
def extraer_visitantes_excel(
    modeladmin,
    request,
    queryset
):

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

    response["Content-Disposition"] = (
        'attachment; filename="historial_visitantes_boccherini.xlsx"'
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Historial"

    encabezados = [
        "ID VISITA",
        "NOMBRES",
        "APELLIDOS",
        "TIPO DOC",
        "DOCUMENTO",
        "TIPO VISITANTE",
        "EMPRESA",
        "PERSONA A VISITAR",
        "DETALLE ADICIONAL",
        "CELULAR",
        "CONTACTO EMERGENCIA",
        "INGRESO",
        "SALIDA",
        "ESTADO",
        "FOTO",
    ]

    for col_num, encabezado in enumerate(
        encabezados,
        1
    ):
        celda = ws.cell(
            row=1,
            column=col_num
        )

        celda.value = encabezado
        celda.font = Font(
            bold=True
        )

    fila = 2

    # IMPORTANTE:
    # Desde Visitor seleccionamos los visitantes,
    # pero exportamos TODAS sus visitas históricas.
    # En el caso de VisitAdmin, el queryset ya son Visitas.
    # Manejamos ambos casos de forma dinámica:
    
    if queryset.model == Visitor:
        visitas = (
            Visit.objects
            .filter(
                visitor__in=queryset
            )
            .select_related("visitor")
            .order_by("-entry_time")
        )
    else:
        # Asume queryset.model == Visit
        visitas = (
            queryset
            .select_related("visitor")
            .order_by("-entry_time")
        )

    for visita in visitas:

        visitante = visita.visitor

        entrada_local = get_bogota_time(
            visita.entry_time
        )

        salida_local = get_bogota_time(
            getattr(
                visita,
                "exit_time",
                None
            )
        )

        tipo_doc = {
            "cedula": "CC",
            "cedula_ciudadania": "CC",
            "ce": "CE",
            "cedula_extranjeria": "CE",
            "pasaporte": "PAS",
        }.get(
            str(
                getattr(
                    visitante,
                    "document_type",
                    ""
                )
            ).lower(),
            ""
        )

        empresa = getattr(
            visitante,
            "company",
            "Particular"
        )

        if (
            visitante
            and getattr(
                visitante,
                "visitor_type",
                ""
            ) == "entrevistado"
        ):
            empresa = "NA"

        datos = [
            visita.id,

            getattr(
                visitante,
                "first_name",
                ""
            ),

            getattr(
                visitante,
                "last_name",
                ""
            ),

            tipo_doc,

            getattr(
                visitante,
                "document_id",
                ""
            ),

            (
                visitante.get_visitor_type_display()
                if hasattr(visitante, "get_visitor_type_display")
                else getattr(visitante, "visitor_type", "")
            ),

            empresa,

            getattr(
                visita,
                "person_to_visit",
                ""
            ),

            getattr(
                visita,
                "reason_detail",
                ""
            ),

            getattr(
                visitante,
                "phone_number",
                ""
            ),

            getattr(
                visitante,
                "emergency_contact",
                ""
            ),

            (
                entrada_local.strftime(
                    "%Y-%m-%d %H:%M"
                )
                if entrada_local
                else ""
            ),

            (
                salida_local.strftime(
                    "%Y-%m-%d %H:%M"
                )
                if salida_local
                else "En Instalaciones"
            ),

            (
                visita.get_status_display()
                if hasattr(
                    visita,
                    "get_status_display"
                )
                else getattr(
                    visita,
                    "status",
                    ""
                )
            ),
        ]

        for col_num, valor in enumerate(
            datos,
            1
        ):
            ws.cell(
                row=fila,
                column=col_num,
                value=valor
            )

        # =====================================================
        # FOTO DE ESA VISITA
        # =====================================================

        if getattr(
            visita,
            "photo",
            None
        ):
            try:
                img = ExcelImage(
                    visita.photo.path
                )

                img.width = 80
                img.height = 80

                ws.add_image(
                    img,
                    f"O{fila}"
                )

                ws.row_dimensions[
                    fila
                ].height = 65

            except Exception:
                pass

        fila += 1

    # =========================================================
    # ANCHOS
    # =========================================================

    anchos = {
        "A": 12,
        "B": 20,
        "C": 20,
        "D": 12,
        "E": 18,
        "F": 20,
        "G": 25,
        "H": 25,
        "I": 45,
        "J": 18,
        "K": 25,
        "L": 20,
        "M": 20,
        "N": 18,
        "O": 18,
    }

    for columna, ancho in anchos.items():
        ws.column_dimensions[
            columna
        ].width = ancho

    wb.save(response)

    return response


# =========================================================
# FILTRO POR PERIODO
# =========================================================
class PeriodoFilter(
    admin.SimpleListFilter
):

    title = "Periodo"
    parameter_name = "periodo"

    def lookups(
        self,
        request,
        model_admin
    ):
        return (
            (
                "semana",
                "Última semana"
            ),
            (
                "mes",
                "Último mes"
            ),
        )

    def queryset(
        self,
        request,
        queryset
    ):
        hoy = timezone.now()

        if self.value() == "semana":
            return queryset.filter(
                entry_time__gte=(
                    hoy - timedelta(days=7)
                )
            )

        if self.value() == "mes":
            return queryset.filter(
                entry_time__gte=(
                    hoy - timedelta(days=30)
                )
            )

        return queryset


# =========================================================
# ADMIN VISITANTES
# =========================================================
@admin.register(Visitor)
class VisitorAdmin(
    admin.ModelAdmin
):

    list_display = (
        "first_name",
        "last_name",
        "document_id",
        "phone_number",
        "emergency_contact_name",
        "emergency_contact_relationship",
        "emergency_contact_phone",
        "visitor_type",
    )

    list_filter = (
        "visitor_type",
    )

    search_fields = (
        "document_id",
        "last_name",
        "first_name",
    )

    actions = [
        extraer_visitantes_excel,
    ]


# =========================================================
# ADMIN VISITAS
# =========================================================
@admin.register(Visit)
class VisitAdmin(
    admin.ModelAdmin
):

    list_display = (
        "id",
        "visitor",
        "person_to_visit",
        "entry_time",
        "status",
    )

    list_filter = (
        PeriodoFilter,
        "visitor__visitor_type",
        "status",
        "entry_time",
    )

    search_fields = (
        "visitor__first_name",
        "visitor__last_name",
        "person_to_visit",
    )

    actions = [
        extraer_historial_csv,
        extraer_visitantes_excel,
    ]
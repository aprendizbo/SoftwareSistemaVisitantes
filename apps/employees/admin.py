import csv
import pytz
from django.contrib import admin
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font
from .models import Employee, EmployeePermission, EarlyDeparture

# Función de conversión forzada a Bogotá
def get_bogota_time(dt):
    if not dt:
        return None
    # Si es "naive", forzamos UTC antes de convertir
    if timezone.is_naive(dt):
        dt = timezone.make_aware(dt, timezone.utc)
    return dt.astimezone(pytz.timezone('America/Bogota'))


@admin.action(description="Extraer Historial (CSV - Hora Colombia)")
def extraer_permisos_csv(modeladmin, request, queryset):
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename="permisos_empleados_boccherini.csv"'
    
    writer = csv.writer(response, delimiter=';')
    # Encabezados
    writer.writerow([
        'ID PERMISO', 'EMPLEADO', 'TIPO DE PERMISO', 
        'HORA SALIDA', 'HORA REGRESO', 'ESTADO', 'DETALLE ADICIONAL'
    ])
    
    for p in queryset.select_related('employee'):
        # Convertimos usando la función forzada
        salida = get_bogota_time(p.departure_time)
        regreso = get_bogota_time(p.return_time)
        
        writer.writerow([
            p.id,
            f"{p.employee.first_name} {p.employee.last_name}" if p.employee else "N/A",
            p.get_permit_type_display(),
            # Usamos formato YYYY-MM-DD HH:MM para que Excel lo interprete como fecha
            salida.strftime('%Y-%m-%d %H:%M') if salida else '',
            regreso.strftime('%Y-%m-%d %H:%M') if regreso else 'Pendiente',
            p.status,
            p.detalle_adicional if p.detalle_adicional else ''
        ])
    return response


@admin.action(description="Extraer Historial (Excel con Fotos)")
def extraer_permisos_excel(modeladmin, request, queryset):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        'attachment; filename="permisos_empleados_boccherini.xlsx"'
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Permisos"

    # Encabezados ajustados (EMPLEADO es una sola columna)
    encabezados = [
        'ID',
        'EMPLEADO',
        'TIPO DOC',
        'DOCUMENTO',
        'ÁREA',
        'TIPO DE PERMISO',
        'HORA SALIDA',
        'HORA REGRESO',
        'ESTADO',
        'DETALLE ADICIONAL',
        'FOTO'
    ]

    for col_num, encabezado in enumerate(encabezados, 1):
        celda = ws.cell(row=1, column=col_num)
        celda.value = encabezado
        celda.font = Font(bold=True)

    fila = 2

    for p in queryset.select_related('employee'):
        salida = get_bogota_time(p.departure_time)
        regreso = get_bogota_time(p.return_time)

        # Extracción segura con nombre completo concatenado
        nombre = f"{p.employee.first_name} {p.employee.last_name}" if p.employee else ""
        tipo_doc = getattr(p.employee, 'document_type', 'N/A')
        documento = getattr(p.employee, 'employee_id', 'N/A')
        area = getattr(p.employee, 'area', 'N/A')

        # Bloque de datos optimizado y corregido
        datos = [
            p.id,
            nombre,
            tipo_doc,
            documento,
            area,
            p.get_permit_type_display(),
            salida.strftime('%Y-%m-%d %H:%M') if salida else '',
            regreso.strftime('%Y-%m-%d %H:%M') if regreso else 'Pendiente',
            p.status,
            p.detalle_adicional if p.detalle_adicional else ''
        ]

        for col_num, valor in enumerate(datos, 1):
            ws.cell(row=fila, column=col_num, value=valor)

        # La foto se inserta de forma correcta en la columna K (columna 11)
        if getattr(p, 'photo', None):
            try:
                img = ExcelImage(p.photo.path)
                img.width = 80
                img.height = 80
                ws.add_image(img, f'K{fila}')
                ws.row_dimensions[fila].height = 65
            except Exception:
                pass

        fila += 1

    # Ajuste preciso de las dimensiones para las 11 columnas (A hasta K)
    ws.column_dimensions['A'].width = 10  # ID
    ws.column_dimensions['B'].width = 35  # EMPLEADO
    ws.column_dimensions['C'].width = 12  # TIPO DOC
    ws.column_dimensions['D'].width = 18  # DOCUMENTO
    ws.column_dimensions['E'].width = 25  # ÁREA
    ws.column_dimensions['F'].width = 25  # TIPO DE PERMISO
    ws.column_dimensions['G'].width = 20  # HORA SALIDA
    ws.column_dimensions['H'].width = 20  # HORA REGRESO
    ws.column_dimensions['I'].width = 15  # ESTADO
    ws.column_dimensions['J'].width = 40  # DETALLE ADICIONAL
    ws.column_dimensions['K'].width = 18  # FOTO

    wb.save(response)
    return response


@admin.action(description="Extraer Historial (Excel con Fotos)")
def extraer_salidas_tempranas_excel(modeladmin, request, queryset):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        'attachment; filename="salidas_tempranas_boccherini.xlsx"'
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Salidas Tempranas"

    encabezados = [
        'ID',
        'NOMBRES',
        'APELLIDOS',
        'TIPO DOC',
        'DOCUMENTO',
        'ÁREA',
        'EMPRESA',
        'FECHA SALIDA',
        'HORA SALIDA',
        'AUTORIZADO POR',
        'DETALLE / JUSTIFICACIÓN',
        'CORREO NOTIFICAR',
        'USUARIO REGISTRO',
        'FECHA REGISTRO',
        'FOTO',
    ]

    for col_num, encabezado in enumerate(encabezados, 1):
        celda = ws.cell(row=1, column=col_num)
        celda.value = encabezado
        celda.font = Font(bold=True)

    fila = 2

    for salida in queryset.select_related('employee', 'usuario_registro'):

        usuario = (
            salida.usuario_registro.username
            if getattr(salida, 'usuario_registro', None)
            else ''
        )

        datos = [
            salida.id,
            salida.first_name,
            salida.last_name,
            salida.get_document_type_display() if hasattr(salida, 'get_document_type_display') else getattr(salida, 'document_type', ''),
            salida.document_id,
            salida.area,
            getattr(salida, 'company', ''),
            salida.departure_date.strftime('%Y-%m-%d') if salida.departure_date else '',
            salida.departure_time.strftime('%H:%M') if salida.departure_time else '',
            getattr(salida, 'autorizado_por', ''),
            getattr(salida, 'detail', ''),
            getattr(salida, 'correo_notificar', ''),
            usuario,
            timezone.localtime(salida.fecha_registro).strftime('%Y-%m-%d %H:%M') if getattr(salida, 'fecha_registro', None) else '',
        ]

        for col_num, valor in enumerate(datos, 1):
            ws.cell(
                row=fila,
                column=col_num,
                value=valor
            )

        # FOTO → columna O
        if getattr(salida, 'photo', None):
            try:
                img = ExcelImage(salida.photo.path)
                img.width = 80
                img.height = 80

                ws.add_image(
                    img,
                    f'O{fila}'
                )

                ws.row_dimensions[fila].height = 65

            except Exception:
                pass

        fila += 1

    # Anchos
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 25
    ws.column_dimensions['K'].width = 45
    ws.column_dimensions['L'].width = 30
    ws.column_dimensions['M'].width = 25
    ws.column_dimensions['N'].width = 20
    ws.column_dimensions['O'].width = 18

    wb.save(response)

    return response


@admin.register(Employee)
class EmployeeAdmin(admin.ModelAdmin):
    list_display = (
        'first_name',
        'last_name',
        'employee_id',
        'area'
    )

    search_fields = (
        'first_name',
        'last_name',
        'employee_id',
        'area'
    )


@admin.register(EmployeePermission)
class EmployeePermissionAdmin(admin.ModelAdmin):
    list_display = ('employee', 'permit_type', 'departure_time', 'return_time', 'status', 'detalle_adicional')
    list_filter = ('status', 'permit_type')
    
    search_fields = (
        'employee__first_name',
        'employee__last_name',
        'employee__employee_id',
        'detalle_adicional'
    )
    
    actions = [
        extraer_permisos_csv,
        extraer_permisos_excel,
    ]


@admin.register(EarlyDeparture)
class EarlyDepartureAdmin(admin.ModelAdmin):
    list_display = (
        'first_name',
        'last_name',
        'document_id',
        'area',
        'departure_date',
        'departure_time',
        'correo_notificar',
        'fecha_registro',
    )

    list_filter = (
        'departure_date',
        'area',
        'document_type',
    )

    search_fields = (
        'first_name',
        'last_name',
        'document_id',
        'area',
        'correo_notificar',
        'detail',
    )

    ordering = (
        '-departure_date',
        '-departure_time',
    )

    actions = [
        extraer_salidas_tempranas_excel,
    ]